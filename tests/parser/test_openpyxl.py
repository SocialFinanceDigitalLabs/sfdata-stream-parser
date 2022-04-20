from pathlib import Path
from tempfile import TemporaryDirectory

import pytest
from more_itertools import split_before
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from sfdata_stream_parser import events
from sfdata_stream_parser.checks import type_check
from sfdata_stream_parser.filters.generic import filter_stream
from sfdata_stream_parser.parser.openpyxl import parse_sheets


@pytest.fixture
def sample_workbook():
    wb = Workbook()

    ws = wb.active
    ws.title = "Test Sheet 1"
    for row in range(40):
        ws.append([f"R{row}C{col}" for col in range(600)])
    tab = Table(displayName="TestTable1", ref="A1:E5")
    ws.add_table(tab)

    ws = wb.create_sheet("Test Sheet 2")
    for row in range(20):
        ws.append([f"R{row}C{col}" for col in range(50)])
    tab = Table(displayName="TestTable2", ref="A1:T50")
    ws.add_table(tab)

    ws = wb.create_sheet("Test Sheet 3")
    for row in range(5):
        ws.append([f"R{row}C{col}" for col in range(5)])
    ws.append([])
    ws.append([f"R6C{col}" for col in range(5)])

    return wb


@pytest.fixture
def sample_file(sample_workbook):
    with TemporaryDirectory() as temp_folder:
        dest_folder = Path(temp_folder)
        dest_file = dest_folder / "sample_file.xlsx"
        sample_workbook.save(dest_file)
        yield dest_file
        dest_file.unlink()


def test_read_from_workbook(sample_workbook):
    wb = sample_workbook

    event_stream = list(parse_sheets(wb))
    assert len(event_stream) == 25177

    assert len(list(filter_stream(event_stream, type_check((events.StartContainer, events.EndContainer))))) == 2

    assert len(list(filter_stream(event_stream, type_check(events.StartTable)))) == 3
    assert len(list(filter_stream(event_stream, type_check(events.EndTable)))) == 3

    pre, table1, table2, table3 = split_before(event_stream, lambda x: type(x) == events.StartTable)

    assert len(list(filter_stream(table1, type_check(events.StartRow)))) == 40
    assert len(list(filter_stream(table2, type_check(events.StartRow)))) == 20
    assert len(list(filter_stream(table3, type_check(events.StartRow)))) == 7

    assert len(list(filter_stream(table1, type_check(events.Cell)))) == 40 * 600
    assert len(list(filter_stream(table2, type_check(events.Cell)))) == 20 * 50
    assert len(list(filter_stream(table3, type_check(events.Cell)))) == 5 * 7


def test_read_from_read_only_workbook(sample_file):
    wb = load_workbook(sample_file, read_only=True, data_only=True)
    test_read_from_workbook(wb)


def test_read_from_string(sample_file):
    event_stream = list(parse_sheets(str(sample_file.absolute())))
    assert len(event_stream) == 25177
    assert event_stream[0].name == str(sample_file.absolute())


def test_read_from_stream(sample_file):
    with open(sample_file, "rb") as f:
        event_stream = list(parse_sheets(f))
    assert len(event_stream) == 25177
    assert event_stream[0].name == str(sample_file.absolute())


def test_read_from_sheet(sample_workbook):
    sheet1 = sample_workbook.worksheets[0]
    event_stream = list(parse_sheets(sheet1))

    assert len(list(filter_stream(event_stream, type_check((events.StartContainer, events.EndContainer))))) == 0

    assert len(list(filter_stream(event_stream, type_check(events.StartTable)))) == 1
    assert len(list(filter_stream(event_stream, type_check(events.EndTable)))) == 1

    assert len(list(filter_stream(event_stream, type_check(events.StartRow)))) == 40
    assert len(list(filter_stream(event_stream, type_check(events.Cell)))) == 40 * 600

