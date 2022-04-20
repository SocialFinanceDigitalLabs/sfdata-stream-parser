from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

from sfdata_stream_parser import events


def _parse_sheet(source: Worksheet):
    yield events.StartTable(name=source.title, type="worksheet")
    for row_ix, row in enumerate(source.rows):
        yield events.StartRow(row_index=row_ix)
        for col_ix, cell in enumerate(row):
            props = dict(value=cell.value, column_index=col_ix)
            props['excel_type'] = getattr(cell, 'data_type', None)
            props['excel_location'] = getattr(cell, 'coordinate', None)
            props['excel_number_format'] = getattr(cell, 'number_format', None)
            yield events.Cell(**props)
        yield events.EndRow()
    yield events.EndTable()


def parse_sheets(source, container_name=None):
    if hasattr(source, 'worksheets'):
        yield events.StartContainer(name=container_name)
        yield from parse_sheets(source.worksheets, container_name=container_name)
        yield events.EndContainer()

    elif isinstance(source, list):
        for sheet in source:
            yield from parse_sheets(sheet, container_name=container_name)

    elif isinstance(source, str):
        if container_name is None:
            container_name = source
        yield from parse_sheets(load_workbook(filename=source, read_only=True, data_only=True),
                                container_name=container_name)

    elif hasattr(source, 'read'):
        if container_name is None and hasattr(source, 'name'):
            container_name = source.name
        yield from parse_sheets(load_workbook(source, read_only=True, data_only=True),
                                container_name=container_name)

    elif isinstance(source, (Worksheet, ReadOnlyWorksheet)):
        yield from _parse_sheet(source)

